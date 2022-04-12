#=====================================================================
# Data analyzer with KERAS API to set a deep learning network
#
# Author   :   Omid Sayedian
# Created  :   Thursday, ‎August ‎5, ‎2021, ‏‎7:14:50 PM
#
# openpyxl 3.0.9
# numpy    1.21.2
# python   3.9.9
# 
#=====================================================================
'''
Project   : Master thesis

Written By: Omid Seyedian
Date&Time : Thursday, October 14, 2021
Location  : Safe-house

# openpyxl-3.0.9
# numpy-1.21.2

#=====================================================================
# Libraries
#=====================================================================
import datetime

from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment
import openpyxl.styles
from openpyxl.styles.borders import Side, Border
from openpyxl.workbook.workbook import Workbook
import serial

#=====================================================================
# Variables
#=====================================================================
vref = 3250

print('Dear user,')
counter = input('How many samples do you wish to capture? ')
print('We are going to acquire {} samples.'.format(counter))
counter = int(counter)

counter_copy = counter
stringish_counter = str(counter) + '\r\n'
bytish_counter = bytes(stringish_counter, 'utf-8')

temple = 200
temple_copy = temple
stringish_temple = str(temple) + '\r\n'
bytish_temple = bytes(stringish_temple, 'utf-8')

mater = input('Please enter the Material Under Test (MUT): ')
mater = '_' + str(mater) + '_'

#=====================================================================
# Colors
#=====================================================================
peach = openpyxl.styles.colors.Color(rgb='CCFFFF')
peach_fill = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=peach)

blue = openpyxl.styles.colors.Color(rgb='FFCC99')
blue_fill = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=blue)

lgrey = openpyxl.styles.colors.Color(rgb='F8F8F8')
lgrey_fill = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=lgrey)

grey = openpyxl.styles.colors.Color(rgb='C0C0C0')
grey_fill = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=grey)

red = openpyxl.styles.colors.Color(rgb='FF0000')
red_fill = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=red)

lgreen = openpyxl.styles.colors.Color(rgb='99FF99')
green_fill = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=lgreen)

yellow = openpyxl.styles.colors.Color(rgb='FFFF66')
yellow_fill = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=yellow)

#=====================================================================
# Functions
#=====================================================================



def make_bold(work_sheet, cell_range):
    bold_font = Font(size=12, bold=True)
    for row in work_sheet[cell_range]:
        for cell in row:
            cell.font = bold_font



def cntr_algn(work_sheet, cell_range):
    align = Alignment(horizontal='center',
                      vertical='center',
                      wrapText=True)
    for row in work_sheet[cell_range]:
        for cell in row:
            cell.alignment = align



def make_data_analog(stng):
    MSB_hex = stng[:2]
    LSB_hex = stng[3:5]

    MSB_hex = str(MSB_hex)
    LSB_hex = str(LSB_hex)

    MSB = int(MSB_hex, 16)
    LSB = int(LSB_hex, 16)
    integral_value = MSB * 256 + LSB
    analog_value = round(integral_value / 32767 * 2048, 3)
    return analog_value



def make_temperature_analog(stng, vref):
    temperature_hex = stng[:2]
    temperature_hex = str(temperature_hex)

    temperature = int(temperature_hex, 16)

    analog_value = round(temperature / 1024 * vref / 10, 1)
    return analog_value



def make_bus_analog(stng, vref):
    bus_hex = stng[3:6]
    bus_hex = str(bus_hex)

    bus = int(bus_hex, 16)

    analog_value = round(bus / 1024 * vref, 0)
    return analog_value



def set_border(work_sheet, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in work_sheet[cell_range]:
        for cell in row:
            cell.border = Border(top=thin,
                                 left=thin,
                                 right=thin,
                                 bottom=thin)



#=====================================================================
# Main Procedure
#=====================================================================
serialPort = serial.Serial(port="COM5",
                           baudrate=19200,
                           bytesize=8,
                           timeout=2,
                           stopbits=serial.STOPBITS_ONE)

#---------------------------------------------------------------Temple
serialString = ""  # Used to hold data coming over UART

serialPort.write(bytish_temple)

raw_temple = []

while(1):
    if(serialPort.in_waiting > 0):
        serialString = serialPort.readline()
        raw_temple.append(serialString.decode('Ascii'))
        # print(serialString.decode('Ascii'))
        # serialPort.write(b"Thank you for sending data \r\n")
        temple_copy -= 1
    if(temple_copy == 0):
        break

#-------------------------------------------------------------- Bridge Defection
serialString = ""  # Used to hold data coming over UART

serialPort.write(bytish_counter)
time_start = datetime.datetime.now()

raw_data = []

while(1):
    if(serialPort.in_waiting > 0):
        serialString = serialPort.readline()
        raw_data.append(serialString.decode('Ascii'))
        # print(serialString.decode('Ascii'))
        # serialPort.write(b"Thank you for sending data \r\n")
        counter_copy -= 1
    if(counter_copy == 0):
        time_stop = datetime.datetime.now()
        break

time_taken = time_stop - time_start
print(time_taken)

#=====================================================================
# Data Processing
#=====================================================================
#-------------------------------- Processing Temperature & Bus Voltage
# print(raw_temple)
# print(len(raw_temple))

analog_temple_temperature = []
analog_temple_bus = []

for i in range(temple):
    analog_temple_bus.append(make_bus_analog(raw_temple[i], vref))

analog_temple_bus_average = sum(analog_temple_bus) / len(analog_temple_bus)

for i in range(temple):
    analog_temple_temperature.append(make_temperature_analog(raw_temple[i], vref))

analog_temple_temperature_average = \
round(sum(analog_temple_temperature) / temple, 1)

# print(analog_temple_temperature_average)
# print(analog_temple_bus_average)
#-------------------------------------------------------Re-calibrating
#------------------------------------------------------Processing Data
# print(raw_data)
# print(len(raw_data))

analog_data = []

for i in range(1, len(raw_data)):
    analog_data.append(make_data_analog(raw_data[i]))

# print(analog_data)
analog_len = len(analog_data)
# print(analog_len)

timer_value = []
value = 0
for i in range(analog_len):
    value = round(value, 5)
    stringish_value = str(value)
    timer_value.append(stringish_value)
    value += 9.984

#=====================================================================
# Create Excel Workbook
#=====================================================================
stringish_time_start = str(time_start)[:19]
stringish_time_taken = str(time_taken)
stringish_time_stop = str(time_stop)[:19]
stringish_analog_len = str(analog_len)

workbook = Workbook()
present_sheet = workbook.active
workbook.create_sheet('Main Sheet')
workbook.create_sheet('Results')
workbook.remove(workbook[workbook.sheetnames[0]])
present_sheet = workbook['Main Sheet']

valid_stringish_date = str(datetime.datetime.now())[:19]
valid_stringish_date = valid_stringish_date.replace(":", "_")
valid_stringish_date = valid_stringish_date.replace("-", "_")
vsd = valid_stringish_date.replace(" ", "_")

filename = "TestReports\\TPS_Method_Analysis" + mater + vsd + ".xlsx"

workbook.save(filename)

present_sheet.cell(row=1, column=1).value = 'Test Name:'
present_sheet.cell(row=1, column=2).value = 'TPS Method Analysis'

present_sheet.cell(row=2, column=1).value = 'Start Time:'
present_sheet.cell(row=2, column=2).value = stringish_time_start

present_sheet.cell(row=3, column=1).value = 'Stop Time:'
present_sheet.cell(row=3, column=2).value = stringish_time_stop

present_sheet.cell(row=4, column=1).value = 'Taken Time:'
present_sheet.cell(row=4, column=2).value = stringish_time_taken

present_sheet.cell(row=5, column=1).value = 'Test Engineer:'
present_sheet.cell(row=5, column=2).value = 'Omid Sayedian'

present_sheet.cell(row=6, column=1).value = 'Firmware Version:'
present_sheet.cell(row=6, column=2).value = 'Ver11'

present_sheet.cell(row=7, column=1).value = 'Application Version:'
present_sheet.cell(row=7, column=2).value = 'Ver2'

present_sheet.cell(row=8, column=1).value = 'Total Samples:'
present_sheet.cell(row=8, column=2).value = stringish_analog_len

present_sheet.cell(row=11, column=1).value = 'Time Elapsed'
present_sheet.cell(row=11, column=2).value = 'Bridge Defection'

present_sheet.cell(row=9, column=1).value = 'Ref. Voltage [mV]'
present_sheet.cell(row=9, column=2).value = analog_temple_bus_average

present_sheet.cell(row=10, column=1).value = 'Amb. Temperature [°C]'
present_sheet.cell(row=10, column=2).value = analog_temple_temperature_average

for x in range(analog_len):
    present_sheet.cell(row=x + 12, column=1).value = timer_value[x]

for x in range(analog_len):
    present_sheet.cell(row=x + 12, column=2).value = analog_data[x]

workbook.save(filename)

#=====================================================================
# Charts and Colors
#=====================================================================
# present_sheet.freeze_panes = 'C10'
cntr_algn(present_sheet, 'A1:C{}'.format(analog_len + 11))
make_bold(present_sheet, 'A1:C11')
set_border(present_sheet, 'A1:B{}'.format(analog_len + 11))

present_sheet.column_dimensions['A'].width = 24
present_sheet.column_dimensions['B'].width = 24

workbook.save(filename)

for cell in present_sheet.iter_rows(min_row=1,
                                    max_row=10,
                                    min_col=1,
                                    max_col=2):
    for j in range(len(cell)):
        cell[j].fill = green_fill

for cell in present_sheet.iter_rows(min_row=11,
                                    max_row=11,
                                    min_col=1,
                                    max_col=2):
    for j in range(len(cell)):
        cell[j].fill = peach_fill

for cell in present_sheet.iter_rows(min_row=12,
                                    max_row=analog_len + 11,
                                    min_col=1,
                                    max_col=2):
    for j in range(len(cell)):
        cell[j].fill = lgrey_fill

chart = LineChart()

data = Reference(worksheet=present_sheet,
                          min_row=11,
                          max_row=analog_len + 11,
                          min_col=2,
                          max_col=2)
chart.add_data(data, titles_from_data=True)

cats = Reference(worksheet=present_sheet,
                  min_row=12,
                  max_row=analog_len + 12,
                  min_col=1,
                  max_col=1)
chart.set_categories(cats)

chart.x_axis.title = "Sample Time [ms]"
chart.y_axis.title = "Bridge Defection [mV]"

chart.width = 28
chart.height = 16

present_sheet.add_chart(chart, "D1")

workbook.save(filename)
print('It is finished!')
