#===============================================================================
# This is my first encounter with KERAS API to set a deep learning network
# openpyxl-3.0.9
# numpy-1.21.2
# keras-2.6.0
# scipy-1.7.1
#===============================================================================
'''
Based on  : Data in "New_Comparison1000.xlsx" file
Project   : Master thesis

Written By: Omid Seyedian
Date&Time : Thursday, October 14, 2021
Location  : Safe-house

# openpyxl-3.0.9
# numpy-1.21.2
# keras-2.6.0
# scipy-1.7.1

Special Tanks To "Eclipse IDE for Eclipse Committers - 2021-06"
'''
#===============================================================================
# Libraries
#===============================================================================
from keras.layers import Dense
from keras.models import Sequential
from keras.utils.np_utils import to_categorical
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import openpyxl.styles
from openpyxl.styles.borders import Border, Side

import numpy as np

#===============================================================================
# Colors
#===============================================================================
peach = openpyxl.styles.colors.Color(rgb='CCFFFF')
peach_fill = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=peach)

blue = openpyxl.styles.colors.Color(rgb='FFCC99')
blue_fill = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=blue)

lgrey = openpyxl.styles.colors.Color(rgb='F8F8F8')
lgrey_fill = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=lgrey)

grey = openpyxl.styles.colors.Color(rgb='C0C0C0')
grey_fill = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=grey)

red = openpyxl.styles.colors.Color(rgb='FF0000')
red_fill = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=red)

lgreen = openpyxl.styles.colors.Color(rgb='99FF99')
green_fill = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=lgreen)

yellow = openpyxl.styles.colors.Color(rgb='FFFF66')
yellow_fill = openpyxl.styles.fills.PatternFill\
            (patternType='solid', fgColor=yellow)

#===============================================================================
# Functions
#===============================================================================





def make_bold(work_sheet, cell_range):
    bold_font = Font(size=12, bold=True)
    for row in work_sheet[cell_range]:
        for cell in row:
            cell.font = bold_font





def cntr_algn(work_sheet, cell_range):
    align = Alignment(horizontal='center',
                      vertical='center',
                      wrapText=True)
    for row in work_sheet[cell_range]:
        for cell in row:
            cell.alignment = align





def make_data_analog(stng):
    MSB_hex = stng[:2]
    LSB_hex = stng[3:5]

    MSB_hex = str(MSB_hex)
    LSB_hex = str(LSB_hex)

    MSB = int(MSB_hex, 16)
    LSB = int(LSB_hex, 16)
    integral_value = MSB * 256 + LSB
    analog_value = round(integral_value / 32767 * 2048, 3)
    return analog_value





def set_border(work_sheet, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in work_sheet[cell_range]:
        for cell in row:
            cell.border = Border(top=thin,
                                 left=thin,
                                 right=thin,
                                 bottom=thin)





def find_matt(integer):
    if integer == 1:
        return  "Air"
    elif integer == 8:
        return  "Aluminium"
    elif integer == 3:
        return  "Brick"
    elif integer == 9:
        return  "Copper"
    elif integer == 4:
        return  "Glass"
    elif integer == 5:
        return  "Ice"
    elif integer == 6:
        return  "Iron"
    elif integer == 2:
        return  "Pine Wood"
    elif integer == 7:
        return  "1%CarbonSteel"
    else:
        return  None





def find_quant(integer):
    if integer == 1:
        return  [0.024, 19]
    elif integer == 8:
        return  [205, 97]
    elif integer == 3:
        return  [0.6, 0.27]
    elif integer == 9:
        return  [385, 111]
    elif integer == 4:
        return  [0.8, 0.34]
    elif integer == 5:
        return  [1.6, 1]
    elif integer == 6:
        return  [80, 23]
    elif integer == 2:
        return  [0.12, 0.082]
    elif integer == 7:
        return  [50, 12]
    else:
        return  None





#===============================================================================
# Loading Training Data
#===============================================================================
file_dir = 'C:\\Users\\ms110930\\Desktop\\'
file_name = 'New_Comparison1000.xlsx'
file = file_dir + file_name

workbook = load_workbook(file,
                         read_only=False,
                         data_only=True)

data_training_sheet = workbook[workbook.sheetnames[0]]

dataset_training = []

for clmn in range(0, 135):
    data = []
    for rw in range(0, 1001):
        data.append(float(data_training_sheet.cell(row=rw + 13,
                                                  column=clmn + 2).value))
    dataset_training.append(data)

dataset_training_matrix = np.array(dataset_training)    # @UndefinedVariable

X_training = dataset_training_matrix[:, 0:1000]
y_training = dataset_training_matrix[:, 1000]

X_training = X_training / 1000
y_training = y_training.astype(int)
y_training = to_categorical(y_training)

#===============================================================================
# Loading Testing Data
#===============================================================================
data_testing_sheet = workbook[workbook.sheetnames[1]]

dataset_testing = []

for clmn in range(0, 45):
    data = []
    for rw in range(0, 1001):
        data.append(float(data_testing_sheet.cell(row=rw + 13,
                                                  column=clmn + 2).value))
    dataset_testing.append(data)

dataset_testing_matrix = np.array(dataset_testing)    # @UndefinedVariable

X_testing = dataset_testing_matrix[:, 0:1000]
y_testing = dataset_testing_matrix[:, 1000]

X_testing = X_testing / 1000
# y_testing = y_testing.astype(int)
# y_testing = to_categorical(y_testing)
y_testing = y_testing.tolist()
test_outcomes = set(y_testing)

#===============================================================================
# Define Neural Network Model
#===============================================================================
model = Sequential()
model.add(Dense(400, input_dim=1000, activation='relu'))
model.add(Dense(800, activation='relu'))
model.add(Dense(200, activation='relu'))
model.add(Dense(10, activation='softmax'))

#===============================================================================
# Compiling The Network
#===============================================================================
model.compile(loss='categorical_crossentropy',
              optimizer='adam',
              metrics=['accuracy'])

#===============================================================================
# Fitting
#===============================================================================
prescaler = 2 ** 5

epothes = 1024*32
batches = 1024*8

for i in range(10):
    model.fit(X_training,
              y_training,
              epochs=int(epothes),
              batch_size=int(batches),
              verbose=0)

    _, accuracy = model.evaluate(X_training, y_training)
    if accuracy * 100 > 90:
        break
print('Accuracy: %.2f' % (accuracy * 100))

#===============================================================================
# Testing
#===============================================================================
predictions = model.predict(X_testing)
print('%d\n' % len(predictions))
# print('%s => %d (expected %d)' % (X_testing[i].tolist(),
#                                   predictions[i],
#                                   y_testing[i]))

X_testing = X_testing.tolist()
for i in range(len(X_testing)):
    for j in range(len(predictions)):
        X_testing[i][j] = round(X_testing[i][j] * 1000, 1)

for i in range(len(predictions)):
    for j in test_outcomes:
        if predictions[i][int(j)] == max(predictions[i]):
            print('%d  =>  Probability: %2.1f' % (j, max(predictions[i])))

#===============================================================================
# Evaluation
#===============================================================================
# scores = model.evaluate(X_testing, y_testing, verbose=0)
# print("%s: %.2f%%" % (model.metrics_names[1], scores[1] * 100))

#===============================================================================
# Create JSON, YAML And HDF5 File
#===============================================================================
model_json = model.to_json()
with open("model.json", "w") as json_file:
    json_file.write(model_json)

model.save_weights("model.h5")
print("Saved model to disk")

# model_yaml = model.to_yaml()
# with open("model.yaml", "w") as yaml_file:
#     yaml_file.write(model_yaml)

#===============================================================================
# Saving Result into Excel
#===============================================================================
'''
This is is one of the most important features of our script.
This process, deletes old file's sheets first.
(Except the main sheets, of course.)'''

for sheet in workbook.sheetnames:
    if sheet not in workbook.sheetnames[0:2]:
        workbook.remove(workbook[sheet])

workbook.create_sheet("Results")
present_sheet = workbook["Results"]
testing_sheet = workbook["Data_Testing"]

present_sheet.cell(row=1, column=1).value = 'Test No'
present_sheet.cell(row=1, column=2).value = 'Group'
present_sheet.cell(row=1, column=3).value = 'Material'
present_sheet.cell(row=1, column=4).value = 'Detected As'
present_sheet.cell(row=1, column=5).value = 'Probability'
present_sheet.cell(row=1, column=6).value = 'Lambda'
present_sheet.cell(row=1, column=7).value = 'Alpha (*E-6)'

probabilities = []
correct_guess = 0

for i in range(len(predictions)):
    for j in test_outcomes:
        if predictions[i][int(j)] == max(predictions[i]):
            present_sheet.cell(row=i + 2, column=1).value = i + 1

            present_sheet.cell(row=i + 2, column=2).value = \
            testing_sheet.cell(row=1013, column=i + 2).value

            present_sheet.cell(row=i + 2, column=3).value = \
            find_matt(testing_sheet.cell(row=1013, column=i + 2).value)

            present_sheet.cell(row=i + 2, column=4).value = find_matt(int(j))

            present_sheet.cell(row=i + 2, column=5).value = max(predictions[i])
            probabilities.append(max(predictions[i]))

            present_sheet.cell(row=i + 2, column=6).value = find_quant(int(j))[0]

            present_sheet.cell(row=i + 2, column=7).value = find_quant(int(j))[1]

            if str(present_sheet.cell(row=i + 2, column=4).value) == \
            str(present_sheet.cell(row=i + 2, column=3).value):
                present_sheet.cell(row=i + 2, column=4).fill = green_fill
                correct_guess = correct_guess + 1
            else:
                present_sheet.cell(row=i + 2, column=4).fill = red_fill

probabilities_average = sum(probabilities) / len(probabilities)
probabilities_minimum = min(probabilities)

print('Avg. Probability: %.8f' % (probabilities_average))
print('Min. Probability: %.8f' % (probabilities_minimum))
print('Correct guesses : %d' % (correct_guess))

cntr_algn(present_sheet, 'A1:G46')
make_bold(present_sheet, 'A1:G1')
set_border(present_sheet, 'A1:G46')

present_sheet.freeze_panes = 'B2'

present_sheet.column_dimensions['A'].width = 14
present_sheet.column_dimensions['B'].width = 14
present_sheet.column_dimensions['C'].width = 14
present_sheet.column_dimensions['D'].width = 14
present_sheet.column_dimensions['E'].width = 14
present_sheet.column_dimensions['F'].width = 14
present_sheet.column_dimensions['G'].width = 14

workbook.save(file)
workbook.close()

print("It is finished!")
